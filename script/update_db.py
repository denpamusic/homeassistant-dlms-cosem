#!/usr/bin/env python3
"""Updates FLAG IDs database from DLMS UA."""

from collections.abc import Generator
from difflib import unified_diff
from io import BytesIO
import json
from pathlib import Path
import sys
from typing import Final
from urllib.request import urlretrieve

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

URL: Final = "https://www.dlms.com/srv/lib/Export_Flagids.php"
FILENAME: Final = "dlms_flagids.json"
COL: dict[str, int] = {
    "flag_id": 1,
    "manufacturer": 2,
    "country": 3,
    "region": 4,
}

JSON_PATH = Path(f"custom_components/dlms_cosem/{FILENAME}")
OVERRIDES: dict[str, str] = {
    "KFM": "Shenzhen Kaifa Technology Co., Ltd.",
}


def worksheet_items(ws: Worksheet) -> Generator[tuple[str, str], None, None]:
    """Return flag id and manufacturer tuple from the worksheet."""
    row = 2
    while row < ws.max_row:
        yield (
            ws.cell(row, COL["flag_id"]).value,
            ws.cell(row, COL["manufacturer"]).value,
        )
        row += 1


sys.stdout.write("Updating flag ids...\n")
xlsx_filename, _ = urlretrieve(URL)
with open(xlsx_filename, "rb") as f:
    buffer = BytesIO(f.read())

wb = openpyxl.load_workbook(buffer)
manufacturers = dict(worksheet_items(wb.active))
wb.close()

for flag_id, override in OVERRIDES.items():
    if (manufacturer := manufacturers.get(flag_id, None)) and manufacturer != override:
        manufacturers[flag_id] = override
        sys.stdout.write(f'Replaced "{manufacturer}" with "{override}"\n')


OLD_JSON = JSON_PATH.read_text()
NEW_JSON = json.dumps(manufacturers, indent=2) + "\n"
sys.stdout.writelines(
    unified_diff(
        OLD_JSON.splitlines(keepends=True),
        NEW_JSON.splitlines(keepends=True),
        fromfile=f"{FILENAME}-old",
        tofile=f"{FILENAME}-new",
        n=0,
    )
)
JSON_PATH.write_text(NEW_JSON)
