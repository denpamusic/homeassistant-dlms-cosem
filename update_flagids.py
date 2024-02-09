"""Update flag ids database."""

import json
import os
from urllib.request import urlretrieve

import openpyxl

DLMS_Flagids_XLSX = "https://www.dlms.com/srv/lib/Export_Flagids.php"
XLSX_FILE = "DLMS_Flagids.xlsx"
JSON_FILE = "custom_components/dlms_cosem/dlms_flagids.json"

urlretrieve(DLMS_Flagids_XLSX, XLSX_FILE)
dataframe = openpyxl.load_workbook(XLSX_FILE)

data = {}

dataframe1 = dataframe.active
for row in range(1, dataframe1.max_row):
    col1, col2 = dataframe1.iter_cols(1, 2)
    data[col1[row].value] = col2[row].value

dataframe.close()
os.remove(JSON_FILE)

with open(JSON_FILE, "w") as f:
    f.write(json.dumps(data, indent=2))
    f.write("\n")
